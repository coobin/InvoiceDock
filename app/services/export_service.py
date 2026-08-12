from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image
from pypdf import PageObject, PdfReader, PdfWriter, Transformation
from pypdf.generic import NameObject, RectangleObject

from app.config import get_settings
from app.models import Invoice
from app.services.extractor import render_first_page

A4_WIDTH = 595.28
A4_HEIGHT = 841.89


def _reader_for_invoice(invoice: Invoice) -> PdfReader:
    path = get_settings().upload_dir / invoice.stored_name
    if invoice.mime_type == "application/pdf":
        return PdfReader(str(path))
    if invoice.mime_type.startswith("image/"):
        image = Image.open(path).convert("RGB")
        buffer = BytesIO()
        image.save(buffer, "PDF", resolution=150.0)
        buffer.seek(0)
        return PdfReader(buffer)
    raise ValueError(f"{invoice.original_name} 不是可排版的 PDF 或图片")


def _place_page(canvas: PageObject, source: PageObject, x: float, y: float, width: float, height: float) -> None:
    source_width = float(source.mediabox.width)
    source_height = float(source.mediabox.height)
    if source_width <= 0 or source_height <= 0:
        return
    scale = min(width / source_width, height / source_height)
    target_width = source_width * scale
    target_height = source_height * scale
    tx = x + (width - target_width) / 2
    ty = y + (height - target_height) / 2
    transform = Transformation().scale(scale, scale).translate(tx, ty)
    existing_annotation_count = len(canvas.get("/Annots", []))
    canvas.merge_transformed_page(source, transform, over=True)
    annotations = canvas.get("/Annots", [])
    for annotation_reference in annotations[existing_annotation_count:]:
        annotation = annotation_reference.get_object()
        rectangle = annotation.get("/Rect")
        if rectangle is None or len(rectangle) != 4:
            continue
        first = transform.apply_on((rectangle[0], rectangle[1]))
        second = transform.apply_on((rectangle[2], rectangle[3]))
        annotation[NameObject("/Rect")] = RectangleObject(
            (
                min(first[0], second[0]),
                min(first[1], second[1]),
                max(first[0], second[0]),
                max(first[1], second[1]),
            )
        )


def make_print_pdf(invoices: list[Invoice], per_page: int = 2) -> bytes:
    if per_page not in {1, 2, 4}:
        raise ValueError("每页数量只支持 1、2 或 4")
    if not invoices:
        raise ValueError("至少选择一张发票")
    readers = [(invoice, _reader_for_invoice(invoice)) for invoice in invoices]
    writer = PdfWriter()
    page_width, page_height = (A4_HEIGHT, A4_WIDTH) if per_page == 4 else (A4_WIDTH, A4_HEIGHT)
    if per_page == 1:
        for _invoice, reader in readers:
            for source in reader.pages:
                canvas = PageObject.create_blank_page(width=page_width, height=page_height)
                _place_page(canvas, source, 28, 28, page_width - 56, page_height - 56)
                writer.add_page(canvas)
    else:
        margin = 24.0
        gap = 12.0
        columns = 1 if per_page == 2 else 2
        rows = 2
        slot_width = (page_width - margin * 2 - gap * (columns - 1)) / columns
        slot_height = (page_height - margin * 2 - gap * (rows - 1)) / rows
        for batch_start in range(0, len(readers), per_page):
            canvas = PageObject.create_blank_page(width=page_width, height=page_height)
            for index, (_invoice, reader) in enumerate(readers[batch_start : batch_start + per_page]):
                row = index // columns
                col = index % columns
                x = margin + col * (slot_width + gap)
                y = page_height - margin - (row + 1) * slot_height - row * gap
                _place_page(canvas, reader.pages[0], x, y, slot_width, slot_height)
            writer.add_page(canvas)
    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def make_invoice_workbook(invoices: list[Invoice]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "发票台账"
    headers = [
        "状态", "查验方式", "发票类型", "发票代码", "发票号码", "开票日期", "销售方", "销售方税号",
        "购买方", "购买方税号", "不含税金额", "税额", "价税合计", "分类", "来源", "抬头警示", "原文件名", "入库时间",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="173B5E")
        cell.alignment = Alignment(horizontal="center")
    for item in invoices:
        sheet.append([
            item.status,
            item.verification_method,
            item.invoice_type,
            item.invoice_code,
            item.invoice_number,
            item.invoice_date,
            item.seller_name,
            item.seller_tax_id,
            item.buyer_name,
            item.buyer_tax_id,
            item.amount,
            item.tax_amount,
            item.total_amount,
            item.category,
            item.source,
            item.title_warning,
            item.original_name,
            item.created_at.isoformat(sep=" ", timespec="seconds"),
        ])
    widths = [12, 12, 20, 16, 22, 14, 32, 22, 32, 22, 15, 15, 15, 14, 14, 24, 36, 21]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index) if index <= 26 else f"A{chr(64 + index - 26)}"].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def make_preview(invoice: Invoice) -> Path | None:
    settings = get_settings()
    target = settings.preview_dir / f"{invoice.id}.jpg"
    if target.exists():
        return target
    source = settings.upload_dir / invoice.stored_name
    try:
        if invoice.mime_type == "application/pdf":
            image = render_first_page(source, scale=1.4)
        elif invoice.mime_type.startswith("image/"):
            image = Image.open(source).convert("RGB")
        else:
            return None
        image.thumbnail((1200, 1200))
        target.parent.mkdir(parents=True, exist_ok=True)
        image.save(target, format="JPEG", quality=82, optimize=True)
        return target
    except Exception:
        return None
